import sys, openpyxl, getpass
from openpyxl.styles import Font

tableWB = openpyxl.Workbook()
tableS1 = tableWB.active
tableS1.title = 'Multiplication Table'
n = int(sys.argv[1])

for i in range(n):
	tableS1.cell(row = i + 2, column = 1).value = i + 1
	tableS1.cell(row = i + 2, column = 1).font = Font(bold=True)

for i in range(n):
	tableS1.cell(row = 1, column = i + 2).value = i + 1
	tableS1.cell(row = 1, column = i + 2).font = Font(bold=True)

for i in range(2, n + 2):
	for j in range(2, n + 2):
		tableS1.cell(row = i, column = j).value = (i - 1) * (j - 1)

tableWB.save('Multiplication_Table.xlsx')